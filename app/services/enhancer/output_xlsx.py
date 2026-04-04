"""
VYRA L1 Support API - XLSX Output Generator
==============================================
Orijinal XLSX dosyasını koruyarak iyileştirilmiş metni yeni sheet olarak ekler.

Author: VYRA AI Team
Version: 1.0.0 (v3.3.3)
"""

import io
import tempfile
from typing import List

from app.services.logging_service import log_system_event
from app.services.enhancer.image_helpers import get_section_text

from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from app.services.document_enhancer import EnhancedSection


def apply_to_original_xlsx(
    original_content: bytes,
    sections: List['EnhancedSection'],
    session_id: str,
    file_type: str = ".xlsx"
) -> str:
    """
    Orijinal XLSX dosyasını koruyarak iyileştirilmiş metni yeni sheet olarak ekler.

    Strateji:
    - Orijinal tüm sheet'ler aynen korunur (veri, format, görseller dahil)
    - Her iyileştirilmiş bölüm için '[Enhanced] OriginalSheetAdı' adında yeni sheet eklenir
    - Yeni sheet'te iyileştirilmiş metin satır satır yazılır
    - no_change bölümler atlanır (yeni sheet oluşturulmaz)

    Returns:
        Geçici XLSX dosya yolu
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(io.BytesIO(original_content))

    # İyileştirilmiş bölümleri yeni sheet olarak ekle
    added_count = 0
    for section in sections:
        if section.change_type == "no_change":
            continue

        text = get_section_text(section)
        if not text or not text.strip():
            continue

        # Sheet adı oluştur (Excel max 31 karakter)
        base_name = (section.heading or f"Bolum_{section.section_index + 1}")[:20]
        sheet_name = f"[E] {base_name}"

        # Aynı isimli sheet varsa numara ekle
        counter = 1
        original_sheet_name = sheet_name
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_sheet_name[:27]}_{counter}"
            counter += 1

        ws = wb.create_sheet(title=sheet_name)

        # Başlık satırı
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

        ws.cell(row=1, column=1, value=f"İyileştirilmiş: {section.heading or f'Bölüm {section.section_index + 1}'}")
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")

        # İyileştirme tipi
        ws.cell(row=2, column=1, value=f"Değişiklik: {section.change_type}")
        ws.cell(row=2, column=1).font = Font(name="Calibri", size=10, italic=True, color="666666")

        if section.explanation:
            ws.cell(row=2, column=2, value=section.explanation)
            ws.cell(row=2, column=2).font = Font(name="Calibri", size=10, italic=True, color="666666")

        # İyileştirilmiş metin satırları
        lines = text.split("\n")
        data_font = Font(name="Calibri", size=11)

        for li, line in enumerate(lines, start=4):
            # Pipe ile ayrılmış satırları sütunlara böl
            if " | " in line or "," in line:
                separator = " | " if " | " in line else ","
                parts = line.split(separator)
                for ci, part in enumerate(parts):
                    cell = ws.cell(row=li, column=ci + 1, value=part.strip())
                    cell.font = data_font
            else:
                cell = ws.cell(row=li, column=1, value=line)
                cell.font = data_font

        # Sütun genişliklerini ayarla
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        added_count += 1

    if added_count == 0:
        log_system_event("INFO", "XLSX: Hiç iyileştirme uygulanmadı, orijinal dosya döndürülüyor", "enhancer")

    # Geçici XLSX dosyasına kaydet
    ext = ".xlsx"
    temp_path = tempfile.mktemp(suffix=ext, prefix=f"enhanced_{session_id}_")
    wb.save(temp_path)

    log_system_event(
        "INFO",
        f"Enhanced XLSX: {added_count} iyileştirilmiş sheet eklendi (orijinal korundu)",
        "enhancer"
    )
    return temp_path
