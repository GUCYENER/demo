"""
Excel Document Processor - v3.1
================================
Akıllı header tespiti ile satır bazlı chunking.
Her Excel dosyasında otomatik olarak header satırını bulur.
Enhanced with consistent metadata format (2024 Best Practices)
"""

from pathlib import Path
from typing import BinaryIO, List, Dict, Any, Tuple

from .base import BaseDocumentProcessor


class ExcelProcessor(BaseDocumentProcessor):
    """Excel dosyalarını işleyen processor - Akıllı header tespiti"""
    
    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls']
    PROCESSOR_NAME = "ExcelProcessor"
    
    def extract_text(self, file_path: Path) -> str:
        """Excel dosyasından metin çıkarır"""
        ext = file_path.suffix.lower()
        
        if ext == '.xlsx':
            return self._extract_xlsx(str(file_path))
        elif ext == '.xls':
            return self._extract_xls(str(file_path))
        else:
            raise ValueError(f"Desteklenmeyen Excel formatı: {ext}")
    
    def extract_text_from_bytes(self, file_obj: BinaryIO, file_name: str) -> str:
        """BytesIO'dan Excel metni çıkarır"""
        ext = f".{file_name.rsplit('.', 1)[-1].lower()}" if "." in file_name else ""
        
        if ext == '.xlsx':
            return self._extract_xlsx_bytes(file_obj)
        elif ext == '.xls':
            return self._extract_xls_bytes(file_obj)
        else:
            raise ValueError(f"Desteklenmeyen Excel formatı: {ext}")
    
    def extract_chunks(self, file_obj: BinaryIO = None, file_name: str = None, file_path: Path = None) -> List[Dict[str, Any]]:
        """
        Excel dosyasından satır bazlı chunk'lar çıkarır.
        Akıllı header tespiti ile her satır ayrı bir chunk olarak döner.
        
        API Signature: Diğer processor'larla tutarlı (file_obj, file_name)
        
        Returns:
            List[dict]: [{"text": "...", "metadata": {"heading": "...", "sheet": "...", ...}}, ...]
        """
        if file_path:
            ext = file_path.suffix.lower()
            if ext == '.xlsx':
                from openpyxl import load_workbook
                wb = load_workbook(str(file_path), data_only=True)
            else:
                import xlrd
                wb = xlrd.open_workbook(str(file_path))
                return self._chunks_from_xlrd(wb, file_name=None, file_path=file_path)
        elif file_obj:
            ext = f".{file_name.rsplit('.', 1)[-1].lower()}" if file_name and "." in file_name else ".xlsx"
            if ext == '.xlsx':
                from openpyxl import load_workbook
                wb = load_workbook(file_obj, data_only=True)
            else:
                import xlrd
                content = file_obj.read()
                wb = xlrd.open_workbook(file_contents=content)
                return self._chunks_from_xlrd(wb, file_name=file_name, file_path=None)
        else:
            raise ValueError("file_path veya file_obj gerekli")
        
        return self._chunks_from_openpyxl(wb, file_name=file_name, file_path=file_path)
    
    def _detect_header_row(self, rows: List[List[Any]]) -> Tuple[int, List[str]]:
        """
        Akıllı header satırı tespiti.
        
        Kriterler:
        1. En az 3 dolu hücre olmalı
        2. Değerler kısa olmalı (ortalama < 50 karakter)
        3. Değerler sayısal olmamalı
        4. Değerler benzersiz olmalı
        5. Tipik header kelimeleri içermeli
        6. Uzun açıklama satırları (tek hücrede çok uzun metin) header DEĞİL
        
        Returns:
            (header_row_index, headers_list)
        """
        if not rows:
            return 0, []
        
        # İlk olarak aktif sütun sayısını belirle (en az bir değeri olan sütunlar)
        active_cols = set()
        for row in rows[:20]:  # İlk 20 satıra bak
            for col_idx, cell in enumerate(row):
                if cell is not None and str(cell).strip() and str(cell).lower() not in ['none', 'nan']:
                    active_cols.add(col_idx)
        
        # En fazla ilk 20 aktif sütunu kullan
        max_cols = min(20, max(active_cols) + 1) if active_cols else len(rows[0]) if rows else 0
        
        best_score = -1
        best_row_idx = 0
        best_headers = []
        
        # İlk 15 satırı kontrol et
        for row_idx, row in enumerate(rows[:15]):
            # Sadece aktif sütunları al
            row_values = [str(row[i]).strip() if i < len(row) and row[i] is not None else "" 
                         for i in range(max_cols)]
            
            # Boş olmayan hücre sayısı
            non_empty = [v for v in row_values if v and v.lower() not in ['none', 'nan', '']]
            
            # En az 1 dolu hücre olmalı (tek sütunlu tablolar için)
            if len(non_empty) < 1:
                continue
            
            # Tek hücrede çok uzun metin varsa bu açıklama satırı olabilir, atla
            max_cell_len = max(len(v) for v in row_values) if row_values else 0
            if len(non_empty) == 1 and max_cell_len > 100:
                continue
            
            score = 0
            
            # Kriter 1: Dolu hücre sayısı (0-20 puan)
            # 3-10 arası dolu hücre idealdir
            if 3 <= len(non_empty) <= 10:
                score += 20
            elif len(non_empty) > 10:
                score += 15
            
            # Kriter 2: Ortalama uzunluk - kısa değerler header olma ihtimali yüksek (0-25 puan)
            avg_len = sum(len(v) for v in non_empty) / len(non_empty) if non_empty else 100
            if avg_len < 25:
                score += 25
            elif avg_len < 40:
                score += 20
            elif avg_len < 60:
                score += 10
            
            # Kriter 3: Sayısal olmama (0-15 puan)
            numeric_count = sum(1 for v in non_empty if self._is_numeric(v))
            non_numeric_ratio = 1 - (numeric_count / len(non_empty)) if non_empty else 0
            score += non_numeric_ratio * 15
            
            # Kriter 4: Benzersizlik (0-15 puan)
            unique_ratio = len(set(non_empty)) / len(non_empty) if non_empty else 0
            score += unique_ratio * 15
            
            # Kriter 5: Tipik header kelimeleri içerme (0-25 puan)
            header_keywords = ['ad', 'adı', 'name', 'id', 'tip', 'tipi', 'type', 'tarih', 'date', 
                             'açıklama', 'description', 'durum', 'status', 'kod', 'code', 
                             'rol', 'role', 'yetki', 'permission', 'talep', 'request', 
                             'kategori', 'category', 'bilgi', 'info', 'search', 'seçimi',
                             'komut', 'komutları', 'command', 'commands']
            keyword_matches = sum(1 for v in non_empty 
                                 if any(kw in v.lower() for kw in header_keywords))
            if keyword_matches > 0:
                score += min(keyword_matches * 5, 25)
            
            if score > best_score:
                best_score = score
                best_row_idx = row_idx
                best_headers = row_values
        
        # Boş header'ları "Sütun1", "Sütun2" şeklinde doldur
        final_headers = []
        for i, h in enumerate(best_headers):
            if h and h.lower() not in ['none', 'nan', '']:
                final_headers.append(h)
            else:
                final_headers.append(f"Sütun{i+1}")
        
        return best_row_idx, final_headers
    
    
    def _is_numeric(self, value: str) -> bool:
        """Değerin sayısal olup olmadığını kontrol eder"""
        if not value:
            return False
        # Sayı formatlarını kontrol et
        value = value.replace(',', '.').replace(' ', '').replace('%', '')
        try:
            float(value)
            return True
        except ValueError:
            return False
    
    def _chunks_from_openpyxl(self, wb, file_name: str = None, file_path: Path = None) -> List[Dict[str, Any]]:
        """openpyxl workbook'tan satır bazlı chunk'lar oluşturur - MERGE HÜCRE DESTEKLİ"""
        chunks = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # 🆕 MERGE HÜCRE DESTEĞİ - Merge range'leri için değer haritası oluştur
            merge_value_map = self._build_merge_value_map(sheet)
            
            # Tüm satırları al ve merge değerlerini uygula
            all_rows = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                # Merge değerlerini çözümle
                resolved_row = self._resolve_merged_values(row, row_idx, merge_value_map)
                all_rows.append(resolved_row)
            
            if not all_rows:
                continue
            
            # Akıllı header tespiti
            header_row_idx, headers = self._detect_header_row(all_rows)
            
            # Header satırından sonraki satırları işle
            for row_idx, row_values in enumerate(all_rows[header_row_idx + 1:], start=header_row_idx + 2):
                row_str_values = [str(cell).strip() if cell is not None else "" for cell in row_values]
                
                # Boş satırları atla
                if not any(row_str_values):
                    continue
                
                # Satırı anlamlı formata çevir
                chunk_text = self._format_row_as_text(headers, row_str_values)
                
                if chunk_text:
                    chunk_text = self._fix_turkish_chars(chunk_text)
                    
                    chunks.append({
                        "text": chunk_text,
                        "metadata": {
                            "type": "excel_row",
                            "heading": sheet_name,  # Sheet adı heading olarak kullanılır
                            "sheet": sheet_name,
                            "row": row_idx,
                            "header_row": header_row_idx + 1,
                            "chunk_index": len(chunks),
                            "source": file_name or (file_path.name if file_path else "")
                        }
                    })
        
        return chunks
    
    def _build_merge_value_map(self, sheet) -> Dict[tuple, Any]:
        """
        Merge range'leri için (row, col) -> value haritası oluşturur.
        Her merge range'in sol üst hücresindeki değeri tüm range'e yayar.
        
        Bu sayede:
        - E1:E5 merge ise, E1'deki değer E2, E3, E4, E5'e de uygulanır
        - Alt satırlardaki None değerler merge değeriyle doldurulur
        """
        merge_map = {}
        
        try:
            for merge_range in sheet.merged_cells.ranges:
                # Sol üst hücrenin değerini al
                min_row = merge_range.min_row
                min_col = merge_range.min_col
                value = sheet.cell(min_row, min_col).value
                
                if value is None:
                    continue
                
                # Range içindeki TÜM hücrelere aynı değeri ata
                for row in range(merge_range.min_row, merge_range.max_row + 1):
                    for col in range(merge_range.min_col, merge_range.max_col + 1):
                        merge_map[(row, col)] = value
        except Exception as e:
            # Merge range okuma hatası - devam et ama logla
            import sys
            print(f"[ExcelProcessor] Merge range okuma hatası: {e}", file=sys.stderr)
        
        return merge_map
    
    def _resolve_merged_values(self, row_values: tuple, row_idx: int, merge_map: Dict[tuple, Any]) -> List[Any]:
        """
        Satırdaki None değerleri merge map'ten doldurur.
        
        Args:
            row_values: iter_rows(values_only=True)'dan gelen tuple
            row_idx: 1-indexed satır numarası
            merge_map: (row, col) -> value haritası
        
        Returns:
            Merge değerleri uygulanmış liste
        """
        resolved = []
        for col_idx, value in enumerate(row_values, start=1):
            if value is None:
                # Merge map'te varsa oradan al
                merged_value = merge_map.get((row_idx, col_idx))
                resolved.append(merged_value)
            else:
                resolved.append(value)
        return resolved
    
    def _chunks_from_xlrd(self, wb, file_name: str = None, file_path: Path = None) -> List[Dict[str, Any]]:
        """xlrd workbook'tan satır bazlı chunk'lar oluşturur"""
        chunks = []
        
        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            
            if sheet.nrows < 2:
                continue
            
            # Tüm satırları al
            rows = []
            for row_idx in range(sheet.nrows):
                row = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
                rows.append(row)
            
            # Akıllı header tespiti
            header_row_idx, headers = self._detect_header_row(rows)
            
            # Header satırından sonraki satırları işle
            for row_idx in range(header_row_idx + 1, sheet.nrows):
                row_values = [str(sheet.cell_value(row_idx, col)).strip() 
                             for col in range(sheet.ncols)]
                
                # Boş satırları atla
                if not any(row_values):
                    continue
                
                # Satırı anlamlı formata çevir
                chunk_text = self._format_row_as_text(headers, row_values)
                
                if chunk_text:
                    chunk_text = self._fix_turkish_chars(chunk_text)
                    
                    chunks.append({
                        "text": chunk_text,
                        "metadata": {
                            "type": "excel_row",
                            "heading": sheet.name,  # Sheet adı heading olarak kullanılır
                            "sheet": sheet.name,
                            "row": row_idx + 1,
                            "header_row": header_row_idx + 1,
                            "chunk_index": len(chunks),
                            "source": file_name or (file_path.name if file_path else "")
                        }
                    })
        
        return chunks
    
    def _format_row_as_text(self, headers: List[str], values: List[str]) -> str:
        """
        Satırı okunabilir metin formatına çevirir.
        Boş değerleri atlar, anlamlı bir metin oluşturur.
        """
        lines = []
        for header, value in zip(headers, values):
            # Boş veya anlamsız değerleri atla
            if not value or value.lower() in ['none', 'nan', '', '0', '0.0']:
                continue
            
            # Header temizle
            header_clean = header.strip()
            if header_clean.startswith('Sütun') and header_clean[5:].isdigit():
                # Sütun1, Sütun2 gibi varsayılan header'ları atla
                lines.append(value)
            else:
                lines.append(f"**{header_clean}:** {value}")
        
        return "\n".join(lines) if lines else ""
    
    def _extract_xlsx(self, file_path: str) -> str:
        """openpyxl ile .xlsx dosyasından metin çıkarır (legacy)"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            return self._formatted_text_from_workbook(wb)
        except ImportError:
            raise ImportError("openpyxl kütüphanesi yüklü değil.")
        except Exception as e:
            raise RuntimeError(f"XLSX işleme hatası: {str(e)}")
    
    def _extract_xlsx_bytes(self, file_obj: BinaryIO) -> str:
        """BytesIO'dan .xlsx metin çıkarır (legacy)"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_obj, data_only=True)
            return self._formatted_text_from_workbook(wb)
        except ImportError:
            raise ImportError("openpyxl kütüphanesi yüklü değil.")
        except Exception as e:
            raise RuntimeError(f"XLSX işleme hatası: {str(e)}")
    
    def _formatted_text_from_workbook(self, wb) -> str:
        """Workbook'u okunabilir formata çevirir"""
        chunks = self._chunks_from_openpyxl(wb)
        return "\n\n---\n\n".join([c["text"] for c in chunks])
    
    def _extract_xls(self, file_path: str) -> str:
        """xlrd ile .xls dosyasından metin çıkarır (legacy)"""
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path)
            chunks = self._chunks_from_xlrd(wb)
            return "\n\n---\n\n".join([c["text"] for c in chunks])
        except ImportError:
            raise ImportError("xlrd kütüphanesi yüklü değil.")
        except Exception as e:
            raise RuntimeError(f"XLS işleme hatası: {str(e)}")
    
    def _extract_xls_bytes(self, file_obj: BinaryIO) -> str:
        """BytesIO'dan .xls metin çıkarır (legacy)"""
        try:
            import xlrd
            content = file_obj.read()
            wb = xlrd.open_workbook(file_contents=content)
            chunks = self._chunks_from_xlrd(wb)
            return "\n\n---\n\n".join([c["text"] for c in chunks])
        except ImportError:
            raise ImportError("xlrd kütüphanesi yüklü değil.")
        except Exception as e:
            raise RuntimeError(f"XLS işleme hatası: {str(e)}")
    
    def get_metadata(self, file_path: Path = None, file_name: str = None) -> dict:
        """Excel metadata'sını çıkarır"""
        return {
            "processor": self.PROCESSOR_NAME,
            "file_name": file_name or (file_path.name if file_path else "unknown"),
        }
